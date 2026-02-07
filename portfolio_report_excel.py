"""
portfolio_report_excel.py
─────────────────────────
Professional multi-sheet Excel portfolio report generator.
Mirrors every chart and table from the Streamlit optimizer.

Usage:
    from portfolio_report_excel import generate_excel_report
    generate_excel_report(report_data, "portfolio_report.xlsx")

Dependencies: openpyxl, pandas, numpy
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint, SeriesLabel
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# Brand colours
# ──────────────────────────────────────────────
NAVY_HEX = "003366"
ACCENT_HEX = "E94560"
BLUE_HEX = "0055A4"
MID_BLUE_HEX = "4A90D9"
LIGHT_BG_HEX = "F0F0F5"
WHITE_HEX = "FFFFFF"
GREEN_HEX = "2E7D32"
RED_HEX = "C62828"
GOLD_HEX = "FCA311"
PURPLE_HEX = "9B5DE5"

HEADER_FILL = PatternFill(start_color=NAVY_HEX, end_color=NAVY_HEX, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=LIGHT_BG_HEX, end_color=LIGHT_BG_HEX, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE_HEX, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=NAVY_HEX, size=14)
SECTION_FONT = Font(name="Calibri", bold=True, color=NAVY_HEX, size=12)
BODY_FONT = Font(name="Calibri", size=10, color="222222")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

_HEAT_NEG = "C62828"
_HEAT_ZERO = "FFFFFF"
_HEAT_POS = "003366"


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────
def _apply_header_row(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _apply_body_rows(ws, row_start, row_end, col_start, col_end):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if (r - row_start) % 2 == 1:
                cell.fill = ALT_ROW_FILL


def _auto_column_widths(ws, min_width=12, max_width=30):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[col_letter].width = min(max_len, max_width)


def _write_title(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left")
    return row + 1


def _write_section(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal="left")
    return row + 1


def _lerp_color(val, lo, hi):
    """Linearly interpolate colour: red (lo) -> white (mid) -> navy (hi)."""
    if hi == lo:
        return _HEAT_ZERO
    t = (val - lo) / (hi - lo)

    def _h2r(h):
        return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)

    def _r2h(r, g, b):
        return f"{int(r):02X}{int(g):02X}{int(b):02X}"

    rn, gn, bn = _h2r(_HEAT_NEG)
    rm, gm, bm = _h2r(_HEAT_ZERO)
    rp, gp, bp = _h2r(_HEAT_POS)

    if t < 0.5:
        s = t / 0.5
        r = rn + (rm - rn) * s
        g = gn + (gm - gn) * s
        b = bn + (bm - bn) * s
    else:
        s = (t - 0.5) / 0.5
        r = rm + (rp - rm) * s
        g = gm + (gp - gm) * s
        b = bm + (bp - bm) * s
    return _r2h(r, g, b)


def _write_matrix(ws, row, matrix, fmt="0.00", apply_heat=False):
    """Write a square matrix with row/col labels and optional heat-map fill."""
    labels = list(matrix.columns)
    n = len(labels)

    ws.cell(row=row, column=1, value="")
    for c, lab in enumerate(labels, 2):
        ws.cell(row=row, column=c, value=lab)
    _apply_header_row(ws, row, 1, n + 1)
    row += 1

    if apply_heat:
        all_vals = matrix.values.flatten()
        lo, hi = float(np.nanmin(all_vals)), float(np.nanmax(all_vals))
    else:
        lo, hi = 0, 1

    for i, lab in enumerate(labels):
        ws.cell(row=row, column=1, value=lab)
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=10, color="222222")
        ws.cell(row=row, column=1).border = THIN_BORDER
        for j in range(n):
            val = float(matrix.iloc[i, j])
            cell = ws.cell(row=row, column=j + 2, value=val)
            cell.number_format = fmt
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            cell.font = BODY_FONT
            if apply_heat:
                hx = _lerp_color(val, lo, hi)
                cell.fill = PatternFill(start_color=hx, end_color=hx, fill_type="solid")
                brightness = sum(int(hx[k:k + 2], 16) for k in (0, 2, 4)) / 3
                if brightness < 140:
                    cell.font = Font(name="Calibri", size=10, color=WHITE_HEX)
        row += 1
    return row


# ──────────────────────────────────────────────
# Sheet 1 – Summary
# ──────────────────────────────────────────────
def _build_summary_sheet(wb, data):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = NAVY_HEX

    row = 1
    row = _write_title(ws, row, 1, "Portfolio Report")
    ws.cell(row=row, column=1,
            value=f"Date: {data.get('report_date', datetime.today().strftime('%Y-%m-%d'))}").font = BODY_FONT
    row += 1
    ws.cell(row=row, column=1,
            value=f"Strategy: {data.get('portfolio_label', 'N/A')}").font = BODY_FONT
    row += 2

    # Key Metrics
    row = _write_section(ws, row, 1, "Key Metrics")
    risk = data.get("risk", {})
    currency = data.get("currency", "USD")
    metrics = [
        ("Portfolio Value", f"{data.get('portfolio_value', 0):,.2f} {currency}"),
        ("Annual Volatility", f"{risk.get('Volatility', 0):.2%}"),
        ("Sharpe Ratio", f"{risk.get('Sharpe', 0):.2f}"),
        ("Max Drawdown", f"{risk.get('Max Drawdown', 0):.2%}"),
    ]
    for c, h in enumerate(["Metric", "Value"], 1):
        ws.cell(row=row, column=c, value=h)
    _apply_header_row(ws, row, 1, 2)
    row += 1
    sb = row
    for label, val in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        row += 1
    _apply_body_rows(ws, sb, row - 1, 1, 2)
    row += 1

    # Performance
    perf = data.get("performance", {})
    if perf:
        row = _write_section(ws, row, 1, "Performance")
        for c, h in enumerate(["Period", "Return"], 1):
            ws.cell(row=row, column=c, value=h)
        _apply_header_row(ws, row, 1, 2)
        row += 1
        sb = row
        for period, ret in perf.items():
            ws.cell(row=row, column=1, value=period)
            cell = ws.cell(row=row, column=2, value=ret)
            cell.number_format = "0.00%"
            cell.font = Font(name="Calibri", size=10,
                             color=GREEN_HEX if ret >= 0 else RED_HEX)
            row += 1
        _apply_body_rows(ws, sb, row - 1, 1, 2)
        row += 1

    # Expected Future Value
    horizon = data.get("investment_horizon")
    fv = data.get("expected_future_value")
    if horizon is not None and fv is not None:
        row = _write_section(ws, row, 1, f"Expected Future Value ({horizon}Y Horizon)")
        fv_metrics = [
            ("Investment", f"{data.get('portfolio_value', 0):,.2f} {currency}"),
            (f"Expected Value ({horizon}Y)", f"{fv:,.2f} {currency}"),
            (f"Expected P&L ({horizon}Y)", f"{data.get('expected_pnl_horizon', 0):+,.2f} {currency}"),
        ]
        fv_up = data.get("fv_upper")
        fv_lo = data.get("fv_lower")
        if fv_up is not None:
            fv_metrics.append(("Optimistic (+1\u03c3)", f"{fv_up:,.2f} {currency}"))
        if fv_lo is not None:
            fv_metrics.append(("Pessimistic (\u22121\u03c3)", f"{fv_lo:,.2f} {currency}"))
        for c, h in enumerate(["Metric", "Value"], 1):
            ws.cell(row=row, column=c, value=h)
        _apply_header_row(ws, row, 1, 2)
        row += 1
        sb = row
        for label, val in fv_metrics:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=val)
            row += 1
        _apply_body_rows(ws, sb, row - 1, 1, 2)
        row += 1

    # Commentary
    commentary = data.get("commentary", [])
    if commentary:
        row = _write_section(ws, row, 1, "Commentary")
        for para in commentary:
            ws.cell(row=row, column=1, value=para).font = BODY_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
            row += 1

    _auto_column_widths(ws)


# ──────────────────────────────────────────────
# Sheet 2 – Holdings & Allocation (table + pie + bar)
# ──────────────────────────────────────────────
def _build_holdings_sheet(wb, data):
    ws = wb.create_sheet("Holdings")
    ws.sheet_properties.tabColor = BLUE_HEX

    row = 1
    row = _write_title(ws, row, 1, "Holdings & Allocation")
    row += 1

    weights = data.get("weights", {})
    holdings_df = data.get("holdings_df")
    currency = data.get("currency", "USD")

    if holdings_df is not None and not holdings_df.empty:
        cols = list(holdings_df.columns)
    else:
        cols = ["Ticker", "Weight"]
        holdings_df = pd.DataFrame([{"Ticker": t, "Weight": w} for t, w in weights.items()])

    for c, h in enumerate(cols, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_header_row(ws, row, 1, len(cols))
    row += 1
    sb = row
    for _, hrow in holdings_df.iterrows():
        for c, col_name in enumerate(cols, 1):
            val = hrow[col_name]
            cell = ws.cell(row=row, column=c)
            if col_name == "Weight":
                cell.value = float(val)
                cell.number_format = "0.00%"
            elif col_name in ("Market Value", "P/L"):
                cell.value = float(val)
                cell.number_format = "#,##0.00"
            else:
                cell.value = str(val)
        row += 1
    _apply_body_rows(ws, sb, row - 1, 1, len(cols))

    # Chart data block
    if len(weights) > 0:
        row += 2
        cds = row  # chart data start
        ws.cell(row=row, column=1, value="Asset")
        ws.cell(row=row, column=2, value="Weight")
        row += 1
        for ticker, w in weights.items():
            ws.cell(row=row, column=1, value=ticker)
            ws.cell(row=row, column=2, value=float(w))
            row += 1
        cde = row - 1  # chart data end

        # Pie chart
        pie = PieChart()
        pie.title = "Portfolio Allocation"
        pie.style = 10
        pie.width = 16
        pie.height = 12
        pie.set_categories(Reference(ws, min_col=1, min_row=cds + 1, max_row=cde))
        pie.add_data(Reference(ws, min_col=2, min_row=cds, max_row=cde), titles_from_data=True)
        ws.add_chart(pie, f"E{cds}")

        # Bar chart
        bar = BarChart()
        bar.title = "Allocation"
        bar.style = 10
        bar.width = 16
        bar.height = 12
        bar.y_axis.title = "Weight"
        bar.x_axis.title = "Asset"
        bar.set_categories(Reference(ws, min_col=1, min_row=cds + 1, max_row=cde))
        bar.add_data(Reference(ws, min_col=2, min_row=cds, max_row=cde), titles_from_data=True)
        bar.shape = 4
        ws.add_chart(bar, f"E{cds + 16}")

    _auto_column_widths(ws)


# ──────────────────────────────────────────────
# Sheet 3 – Cumulative Returns (ALL individual assets)
# ──────────────────────────────────────────────
def _build_cumulative_returns_sheet(wb, data):
    ws = wb.create_sheet("Cumulative Returns")
    ws.sheet_properties.tabColor = MID_BLUE_HEX

    row = 1
    row = _write_title(ws, row, 1, "Cumulative Returns (Base = 100)")
    row += 1

    cum_returns = data.get("cum_returns")
    if cum_returns is None or (hasattr(cum_returns, "empty") and cum_returns.empty):
        ws.cell(row=row, column=1, value="No cumulative returns data available.").font = BODY_FONT
        return

    # Downsample if large
    if len(cum_returns) > 520:
        cum_returns = cum_returns.resample("W-FRI").last().dropna()

    assets = list(cum_returns.columns)
    n_assets = len(assets)

    headers = ["Date"] + assets
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_header_row(ws, row, 1, len(headers))
    header_row = row
    row += 1

    sb = row
    for dt, vals in cum_returns.iterrows():
        ws.cell(row=row, column=1, value=dt.date() if hasattr(dt, "date") else dt)
        ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
        for c, asset in enumerate(assets, 2):
            ws.cell(row=row, column=c, value=float(vals[asset])).number_format = "0.00"
        row += 1
    _apply_body_rows(ws, sb, row - 1, 1, len(headers))

    # Line chart — one series per asset
    chart = LineChart()
    chart.title = "Cumulative Returns (Base = 100)"
    chart.style = 10
    chart.width = 30
    chart.height = 16
    chart.y_axis.title = "Growth of 100"
    chart.x_axis.title = "Date"
    chart.x_axis.number_format = "YYYY-MM-DD"
    chart.x_axis.tickLblSkip = max(1, (row - header_row) // 12)

    dates_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=row - 1)
    for i in range(n_assets):
        ref = Reference(ws, min_col=i + 2, min_row=header_row, max_row=row - 1)
        chart.add_data(ref, titles_from_data=True)
    chart.set_categories(dates_ref)
    for s in chart.series:
        s.graphicalProperties.line.width = 18000

    ws.add_chart(chart, f"{get_column_letter(n_assets + 3)}{header_row}")
    _auto_column_widths(ws)


# ──────────────────────────────────────────────
# Sheet 4 – Portfolio vs Benchmark
# ──────────────────────────────────────────────
def _build_performance_sheet(wb, data):
    ws = wb.create_sheet("Portfolio Performance")
    ws.sheet_properties.tabColor = "00B4D8"

    row = 1
    row = _write_title(ws, row, 1, "Portfolio vs Benchmark")
    row += 1

    port_series = data.get("portfolio_series")
    bench_series = data.get("benchmark_series")

    if port_series is None or len(port_series) == 0:
        ws.cell(row=row, column=1, value="No performance data available.").font = BODY_FONT
        return

    has_bench = bench_series is not None and len(bench_series) > 0

    if len(port_series) > 520:
        port_series = port_series.resample("W-FRI").last().dropna()
        if has_bench:
            bench_series = bench_series.resample("W-FRI").last().dropna()

    headers = ["Date", data.get("portfolio_label", "Portfolio")]
    if has_bench:
        headers.append(data.get("benchmark_label", "Benchmark"))
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _apply_header_row(ws, row, 1, len(headers))
    header_row = row
    row += 1

    sb = row
    for i, (dt, val) in enumerate(port_series.items()):
        ws.cell(row=row, column=1, value=dt.date() if hasattr(dt, "date") else dt)
        ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
        ws.cell(row=row, column=2, value=float(val)).number_format = "0.00"
        if has_bench:
            try:
                bval = bench_series.get(dt, None) if hasattr(bench_series, "get") else None
                if bval is None and i < len(bench_series):
                    bval = float(bench_series.iloc[i])
            except Exception:
                bval = None
            if bval is not None:
                ws.cell(row=row, column=3, value=float(bval)).number_format = "0.00"
        row += 1
    _apply_body_rows(ws, sb, row - 1, 1, len(headers))

    chart = LineChart()
    chart.title = "Portfolio vs Benchmark (Base 100)"
    chart.style = 10
    chart.width = 28
    chart.height = 14
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Date"
    chart.x_axis.number_format = "YYYY-MM-DD"
    chart.x_axis.tickLblSkip = max(1, (row - header_row) // 12)

    dates_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=row - 1)
    port_ref = Reference(ws, min_col=2, min_row=header_row, max_row=row - 1)
    chart.add_data(port_ref, titles_from_data=True)
    chart.set_categories(dates_ref)
    chart.series[0].graphicalProperties.line.width = 22000

    if has_bench:
        bench_ref = Reference(ws, min_col=3, min_row=header_row, max_row=row - 1)
        chart.add_data(bench_ref, titles_from_data=True)
        chart.series[1].graphicalProperties.line.width = 22000

    ws.add_chart(chart, f"E{header_row}")
    _auto_column_widths(ws)


# ──────────────────────────────────────────────
# Sheet 5 – Efficient Frontier + CML
# ──────────────────────────────────────────────
def _build_frontier_sheet(wb, data):
    ws = wb.create_sheet("Frontier & CML")
    ws.sheet_properties.tabColor = GOLD_HEX

    row = 1
    row = _write_title(ws, row, 1, "Efficient Frontier & Capital Market Line")
    row += 1

    frontier = data.get("frontier", {})
    pv_cloud = frontier.get("pv_cloud", [])
    pr_cloud = frontier.get("pr_cloud", [])
    sharpe_cloud = frontier.get("sharpe_cloud", [])
    cml_v = frontier.get("cml_v")
    cml_r = frontier.get("cml_r")
    port_v = frontier.get("port_v")
    port_r = frontier.get("port_r")
    selected_label = frontier.get("selected_label", "Selected")

    if not pv_cloud or not pr_cloud:
        ws.cell(row=row, column=1, value="No frontier data available.").font = BODY_FONT
        return

    # ── Cloud data (cols A-B) ──
    row = _write_section(ws, row, 1, "Simulated Portfolios (Frontier Cloud)")
    ws.cell(row=row, column=1, value="Volatility")
    ws.cell(row=row, column=2, value="Return")
    _apply_header_row(ws, row, 1, 2)
    cloud_header = row
    row += 1

    pv_arr = np.array(pv_cloud)
    pr_arr = np.array(pr_cloud)
    sh_arr = np.array(sharpe_cloud) if len(sharpe_cloud) > 0 else np.array([])
    # Downsample cloud for Excel (max 2000 pts)
    max_pts = 2000
    if len(pv_arr) > max_pts:
        idx = np.linspace(0, len(pv_arr) - 1, max_pts, dtype=int)
        pv_arr, pr_arr = pv_arr[idx], pr_arr[idx]
        if len(sh_arr) > 0:
            sh_arr = sh_arr[idx]

    for v, r in zip(pv_arr, pr_arr):
        ws.cell(row=row, column=1, value=float(v))
        ws.cell(row=row, column=2, value=float(r))
        row += 1
    cloud_end = row - 1

    # ── CML data (cols D-E) ──
    has_cml = cml_v is not None and cml_r is not None and len(cml_v) > 0
    cml_end = cloud_header
    if has_cml:
        ws.cell(row=cloud_header, column=4, value="CML Vol")
        ws.cell(row=cloud_header, column=5, value="CML Return")
        _apply_header_row(ws, cloud_header, 4, 5)
        for i, (cv, cr) in enumerate(zip(cml_v, cml_r)):
            ws.cell(row=cloud_header + 1 + i, column=4, value=float(cv))
            ws.cell(row=cloud_header + 1 + i, column=5, value=float(cr))
        cml_end = cloud_header + len(cml_v)

    # ── Selected portfolio marker (cols G-H) ──
    sel_end = cloud_header
    ws.cell(row=cloud_header, column=7, value="Sel Vol")
    ws.cell(row=cloud_header, column=8, value="Sel Return")
    _apply_header_row(ws, cloud_header, 7, 8)
    if port_v is not None and port_r is not None:
        ws.cell(row=cloud_header + 1, column=7, value=float(port_v))
        ws.cell(row=cloud_header + 1, column=8, value=float(port_r))
        sel_end = cloud_header + 1

    # ── Max Sharpe marker (cols J-K) ──
    if len(sh_arr) > 0:
        ws.cell(row=cloud_header, column=10, value="MaxSh Vol")
        ws.cell(row=cloud_header, column=11, value="MaxSh Ret")
        _apply_header_row(ws, cloud_header, 10, 11)
        i_max = int(np.nanargmax(sh_arr))
        ws.cell(row=cloud_header + 1, column=10, value=float(pv_arr[i_max]))
        ws.cell(row=cloud_header + 1, column=11, value=float(pr_arr[i_max]))

    # ── Scatter chart ──
    chart = ScatterChart()
    chart.title = "Frontier (Risky-Only) + CML"
    chart.style = 13
    chart.width = 28
    chart.height = 16
    chart.x_axis.title = "Annualized Volatility"
    chart.y_axis.title = "Annualized Return"

    # Cloud scatter
    xv = Reference(ws, min_col=1, min_row=cloud_header + 1, max_row=cloud_end)
    yv = Reference(ws, min_col=2, min_row=cloud_header + 1, max_row=cloud_end)
    chart.add_data(yv, titles_from_data=False)
    chart.series[0].xvalues = xv
    chart.series[0].tx = SeriesLabel(v="Simulated Portfolios")
    chart.series[0].graphicalProperties.line.noFill = True
    chart.series[0].marker.symbol = "circle"
    chart.series[0].marker.size = 3

    # CML line
    if has_cml:
        xc = Reference(ws, min_col=4, min_row=cloud_header + 1, max_row=cml_end)
        yc = Reference(ws, min_col=5, min_row=cloud_header + 1, max_row=cml_end)
        chart.add_data(yc, titles_from_data=False)
        s = chart.series[-1]
        s.xvalues = xc
        s.tx = SeriesLabel(v="CML (RF + Tangency)")
        s.marker.symbol = "none"
        s.graphicalProperties.line.width = 24000

    # Selected marker
    if port_v is not None:
        xs = Reference(ws, min_col=7, min_row=cloud_header + 1, max_row=sel_end)
        ys = Reference(ws, min_col=8, min_row=cloud_header + 1, max_row=sel_end)
        chart.add_data(ys, titles_from_data=False)
        s = chart.series[-1]
        s.xvalues = xs
        s.tx = SeriesLabel(v=f"Selected ({selected_label})")
        s.graphicalProperties.line.noFill = True
        s.marker.symbol = "diamond"
        s.marker.size = 12

    # Max Sharpe marker
    if len(sh_arr) > 0:
        xm = Reference(ws, min_col=10, min_row=cloud_header + 1, max_row=cloud_header + 1)
        ym = Reference(ws, min_col=11, min_row=cloud_header + 1, max_row=cloud_header + 1)
        chart.add_data(ym, titles_from_data=False)
        s = chart.series[-1]
        s.xvalues = xm
        s.tx = SeriesLabel(v="Max Sharpe (cloud)")
        s.graphicalProperties.line.noFill = True
        s.marker.symbol = "star"
        s.marker.size = 12

    ws.add_chart(chart, f"A{cloud_end + 3}")
    _auto_column_widths(ws)


# ──────────────────────────────────────────────
# Sheet 6 – Correlation & Covariance Matrices
# ──────────────────────────────────────────────
def _build_matrices_sheet(wb, data):
    ws = wb.create_sheet("Correlation & Covariance")
    ws.sheet_properties.tabColor = PURPLE_HEX

    row = 1
    row = _write_title(ws, row, 1, "Correlation & Covariance Matrices")
    row += 1

    corr = data.get("corr_matrix")
    cov_mat = data.get("cov_matrix")

    if corr is not None and not corr.empty:
        row = _write_section(ws, row, 1, "Correlation Matrix")
        row = _write_matrix(ws, row, corr, fmt="0.00", apply_heat=True)
        row += 2

    if cov_mat is not None and not cov_mat.empty:
        row = _write_section(ws, row, 1, "Covariance Matrix (Annualized)")
        row = _write_matrix(ws, row, cov_mat, fmt="0.000000", apply_heat=True)

    _auto_column_widths(ws)


# ──────────────────────────────────────────────
# Sheet 7 – Betas & CAPM
# ──────────────────────────────────────────────
def _build_betas_sheet(wb, data):
    ws = wb.create_sheet("Betas & CAPM")
    ws.sheet_properties.tabColor = ACCENT_HEX

    row = 1
    benchmark = data.get("benchmark_label", "SPY")
    row = _write_title(ws, row, 1, f"Betas & CAPM vs {benchmark}")
    row += 1

    betas = data.get("betas")
    capm_df = data.get("capm_df")

    beta_header = row
    if betas is not None and len(betas) > 0:
        row = _write_section(ws, row, 1, f"Asset Betas vs {benchmark}")
        ws.cell(row=row, column=1, value="Asset")
        ws.cell(row=row, column=2, value=f"Beta vs {benchmark}")
        _apply_header_row(ws, row, 1, 2)
        beta_header = row
        row += 1
        sb = row
        sorted_betas = betas.sort_values(ascending=False)
        for asset, beta_val in sorted_betas.items():
            ws.cell(row=row, column=1, value=str(asset))
            ws.cell(row=row, column=2, value=float(beta_val)).number_format = "0.00"
            row += 1
        beta_end = row - 1
        _apply_body_rows(ws, sb, beta_end, 1, 2)

        # Beta bar chart
        bar = BarChart()
        bar.title = f"Betas vs {benchmark}"
        bar.style = 10
        bar.width = 18
        bar.height = 12
        bar.y_axis.title = "Beta"
        bar.set_categories(Reference(ws, min_col=1, min_row=beta_header + 1, max_row=beta_end))
        bar.add_data(Reference(ws, min_col=2, min_row=beta_header, max_row=beta_end), titles_from_data=True)
        bar.shape = 4
        ws.add_chart(bar, f"D{beta_header}")
        row += 1

    # CAPM table
    if capm_df is not None and not capm_df.empty:
        row = max(row, beta_header + 18) if betas is not None else row
        row = _write_section(ws, row, 1, "CAPM Regression Results")
        capm_display = capm_df.copy()
        if "Alpha (daily)" in capm_display.columns:
            capm_display["Alpha (annualized)"] = (1.0 + capm_display["Alpha (daily)"]) ** 252 - 1.0
            capm_display = capm_display[["Asset", "Alpha (annualized)", "Beta", "R^2", "Corr (excess vs mkt)"]]
        cols = list(capm_display.columns)
        for c, h in enumerate(cols, 1):
            ws.cell(row=row, column=c, value=h)
        _apply_header_row(ws, row, 1, len(cols))
        row += 1
        sb = row
        for _, crow in capm_display.iterrows():
            for c, col_name in enumerate(cols, 1):
                val = crow[col_name]
                cell = ws.cell(row=row, column=c)
                if col_name == "Asset":
                    cell.value = str(val)
                elif col_name == "Alpha (annualized)":
                    cell.value = float(val)
                    cell.number_format = "0.00%"
                else:
                    cell.value = float(val)
                    cell.number_format = "0.00"
            row += 1
        _apply_body_rows(ws, sb, row - 1, 1, len(cols))

    if betas is None and (capm_df is None or capm_df.empty):
        ws.cell(row=row, column=1,
                value="Betas / CAPM not available (benchmark download failed or insufficient overlap).").font = BODY_FONT

    _auto_column_widths(ws)


# ──────────────────────────────────────────────
# Sheet 8 – Risk Analysis
# ──────────────────────────────────────────────
def _build_risk_sheet(wb, data):
    ws = wb.create_sheet("Risk Analysis")
    ws.sheet_properties.tabColor = RED_HEX

    row = 1
    row = _write_title(ws, row, 1, "Risk Analysis")
    row += 1

    risk = data.get("risk", {})
    if risk:
        for c, h in enumerate(["Risk Metric", "Value"], 1):
            ws.cell(row=row, column=c, value=h)
        _apply_header_row(ws, row, 1, 2)
        row += 1
        sb = row
        for metric, val in risk.items():
            if val is None:
                continue
            ws.cell(row=row, column=1, value=metric)
            cell = ws.cell(row=row, column=2)
            if any(k in metric.lower() for k in ("sharpe", "sortino", "ratio", "beta", "r-squared")):
                cell.value = float(val)
                cell.number_format = "0.00"
            else:
                cell.value = float(val)
                cell.number_format = "0.00%"
            row += 1
        _apply_body_rows(ws, sb, row - 1, 1, 2)
    row += 1

    # Daily returns statistics
    port_series = data.get("portfolio_series")
    if port_series is not None and len(port_series) > 1:
        daily_rets = port_series.pct_change().dropna()
        row = _write_section(ws, row, 1, "Daily Returns Statistics")
        stats = [
            ("Mean Daily Return", float(daily_rets.mean()), "0.0000%"),
            ("Std Dev (Daily)", float(daily_rets.std()), "0.0000%"),
            ("Skewness", float(daily_rets.skew()), "0.00"),
            ("Kurtosis", float(daily_rets.kurtosis()), "0.00"),
            ("Min Daily Return", float(daily_rets.min()), "0.0000%"),
            ("Max Daily Return", float(daily_rets.max()), "0.0000%"),
            ("Trading Days", len(daily_rets), "#,##0"),
        ]
        for c, h in enumerate(["Statistic", "Value"], 1):
            ws.cell(row=row, column=c, value=h)
        _apply_header_row(ws, row, 1, 2)
        row += 1
        sb = row
        for label, val, fmt in stats:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2)
            cell.value = int(val) if label == "Trading Days" else val
            cell.number_format = fmt
            row += 1
        _apply_body_rows(ws, sb, row - 1, 1, 2)

    _auto_column_widths(ws)


# ──────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────
def generate_excel_report(report_data, output_path):
    """
    Generate a professional multi-sheet Excel workbook with all optimizer charts.

    Sheets:
        1. Summary             – key metrics, performance, commentary
        2. Holdings            – allocation table + pie chart + bar chart
        3. Cumulative Returns  – all-asset growth-of-100 line chart
        4. Portfolio Performance – portfolio vs benchmark line chart
        5. Frontier & CML      – efficient frontier scatter + CML line
        6. Correlation & Cov   – heat-mapped matrices
        7. Betas & CAPM        – beta bar chart + CAPM regression table
        8. Risk Analysis       – risk metrics + daily return statistics
    """
    wb = Workbook()

    _build_summary_sheet(wb, report_data)
    _build_holdings_sheet(wb, report_data)
    _build_cumulative_returns_sheet(wb, report_data)
    _build_performance_sheet(wb, report_data)
    _build_frontier_sheet(wb, report_data)
    _build_matrices_sheet(wb, report_data)
    _build_betas_sheet(wb, report_data)
    _build_risk_sheet(wb, report_data)

    wb.save(output_path)
    return output_path
